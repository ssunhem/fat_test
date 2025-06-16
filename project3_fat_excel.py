import cv2  # นำเข้าไลบรารี OpenCV สำหรับประมวลผลภาพ
import numpy as np  # นำเข้า NumPy สำหรับการคำนวณทางตัวเลขและเมทริกซ์
import tkinter as tk  # นำเข้า Tkinter สำหรับสร้าง GUI
from tkinter import filedialog, messagebox, Toplevel  # โมดูลย่อยสำหรับเปิดไฟล์, แสดงกล่องข้อความ และสร้างหน้าต่างย่อย
from PIL import Image, ImageTk  # นำเข้า PIL สำหรับจัดการภาพใน Tkinter
import os  # นำเข้าโมดูลระบบไฟล์
from tkinter import ttk  # นำเข้า ttk สำหรับวิดเจ็ตเสริมใน Tkinter
from datetime import datetime  # นำเข้า datetime สำหรับวันที่และเวลา

import openpyxl  # นำเข้า openpyxl สำหรับจัดการไฟล์ Excel
from openpyxl.drawing.image import Image as ExcelImage  # นำเข้า Image จาก openpyxl สำหรับแทรกภาพใน Excel
from io import BytesIO  # นำเข้า BytesIO สำหรับสร้าง buffer ภาพ
from openpyxl.styles import Font, PatternFill  # นำเข้า Font และ PatternFill สำหรับตกแต่งเซลล์ Excel


class FatDetectionApp:
    def __init__(self, root):
        # ตัวสร้าง: กำหนดหน้าต่างหลักและค่าเริ่มต้น
        self.root = root
        self.root.title("Fat & Lean")  # กำหนดชื่อหน้าต่าง
        self.root.geometry("1280x800")  # กำหนดขนาดหน้าต่าง
        
        self.image_panels = [None] * 8  # เก็บ Label สำหรับแสดงภาพ 8 ช่อง
        self.fat_percentages = [None] * 8  # เก็บเปอร์เซ็นต์ไขมันแต่ละภาพ
        self.calculated_status = False  # สถานะว่าคำนวณค่าแล้วหรือไม่

        self.root.configure(bg="black")  # ตั้งพื้นหลังหน้าต่างเป็นสีดำ
        
        # ค้นหาและโหลดไอคอนบริษัทจากไดร์ฟ D
        icon_path = None
        for root_dir, dirs, files in os.walk("D:\\"):
            for file in files:
                if file == "LogoCompanyB.png":  # ไฟล์ไอคอน
                    icon_path = os.path.join(root_dir, file)
                    break
            if icon_path:
                break
        if icon_path:
            icon_img = Image.open(icon_path)
            icon_img = ImageTk.PhotoImage(icon_img)
            self.root.wm_iconphoto(True, icon_img)
        
        self.create_widgets()  # เรียกสร้างวิดเจ็ตทั้งหมด
    
    def create_widgets(self):
        # สร้างส่วนหัว (header) ประกอบด้วยโลโก้และชื่อแอป
        header_frame = tk.Frame(self.root, bg="black")
        header_frame.pack(pady=10)

        # ค้นหาและโหลดภาพโลโก้จากไดร์ฟ D
        logo_path = None
        for root_dir, dirs, files in os.walk("D:\\"):
            for file in files:
                if file == "LogoCompanyB.jpg":  # ไฟล์โลโก้
                    logo_path = os.path.join(root_dir, file)
                    break
            if logo_path:
                break
        if logo_path:
            logo_img_pil = Image.open(logo_path)
            logo_img_pil = logo_img_pil.resize((90, 90))  # ปรับขนาดโลโก้
            self.logo_img = ImageTk.PhotoImage(logo_img_pil)
        else:
            self.logo_img = None

        # แสดงโลโก้ใน header ถ้ามี
        logo_label = tk.Label(header_frame, image=self.logo_img, bg="black")
        logo_label.grid(row=0, column=0, padx=10)

        # แสดงชื่อแอป ข้อความอยู่ใน text= เป็นภาษาอังกฤษ
        title_label = tk.Label(
            header_frame,
            text="Fat Percentage in Ground Beef",
            font=("Arial", 24, "bold"),
            bg="black",
            fg="#CA5C27"
        )
        title_label.grid(row=0, column=1, padx=10)

        # สร้าง frame สำหรับรับข้อมูลผู้ใช้
        input_frame = tk.Frame(self.root, bg="black")
        input_frame.pack()

        self.entries = {}  # เก็บ Entry fields

        # ฟังก์ชัน validate ข้อความไม่เกินความยาวที่กำหนด
        def validate_input(value, max_length):
            return len(value) <= max_length

        # สร้าง command สำหรับ validate
        vcmd_15 = (self.root.register(lambda text: validate_input(text, 15)), "%P")
        vcmd_20 = (self.root.register(lambda text: validate_input(text, 20)), "%P")

        # กำหนดป้ายชื่อและความยาวสูงสุดของ Entry
        labels = [("ID_Manufacturer", 15), ("ID_QC", 15), ("LOT NO.", 20)]

        for i, (label, max_length) in enumerate(labels):
            # สร้าง Label อธิบายช่องกรอก
            tk.Label(
                input_frame,
                text=label,
                font=("Arial", 12, "bold"),
                fg='white',
                bg='black'
            ).grid(row=i, column=0, padx=5, pady=2)
            # สร้าง Entry พร้อม validate
            entry = tk.Entry(
                input_frame,
                validate="key",
                validatecommand=vcmd_15 if max_length == 15 else vcmd_20,
                font=("Arial", 12, 'bold'),
                fg='black'
            )
            entry.grid(row=i, column=1, padx=5, pady=2)
            self.entries[label] = entry  # เก็บ Entry ลง dict

        # แสดงป้าย Date & Time และ Label สำหรับแสดงวันที่
        tk.Label(
            input_frame,
            text="Date & Time",
            font=("Arial", 12, "bold"),
            fg='white',
            bg='black'
        ).grid(row=3, column=0, padx=5, pady=2)

        self.datetime_label = tk.Label(
            input_frame,
            text="",
            font=("Arial", 12, "bold"),
            bg='black',
            fg='white'
        )
        self.datetime_label.grid(row=3, column=1, padx=5, pady=2)
        self.update_datetime()  # เริ่มอัปเดตเวลาปัจจุบันทุกวินาที

        # ฟังก์ชัน validate เปอร์เซ็นต์ (0–100)
        def validate_percentage(value_if_allowed):
            if value_if_allowed == "":
                return True
            if value_if_allowed.isdigit():
                if value_if_allowed.startswith("0") and len(value_if_allowed) > 1:
                    return False
                return 0 <= int(value_if_allowed) <= 100
            return False

        vcmd = (self.root.register(validate_percentage), "%P")

        # ป้ายและ Entry สำหรับ Target Fat Percentage
        tk.Label(
            input_frame,
            text="Target Fat Percentage (+-5%)",
            font=("Arial", 12, "bold"),
            fg='white',
            bg='black'
        ).grid(row=4, column=0, padx=5, pady=2)
        self.target_fat_percentage = tk.Entry(
            input_frame,
            font=("Arial", 12, "bold"),
            justify="center",
            width=17,
            validate="key",
            validatecommand=vcmd
        )
        self.target_fat_percentage.grid(row=4, column=1, padx=5, pady=2)
        # แสดงสัญลักษณ์ %
        tk.Label(
            input_frame,
            text="%",
            font=("Arial", 12, "bold"),
            fg='white',
            bg='black'
        ).grid(row=4, column=2, padx=(3, 5), pady=2)

        # สร้างกรอบสำหรับแสดงภาพ 8 รูป
        self.frame = tk.Frame(self.root, bg='black')
        self.frame.pack(pady=10)

        self.image_boxes = []  # เก็บ Frame แต่ละช่องภาพ
        self.delete_buttons = []  # เก็บปุ่มลบภาพ
        
        for i in range(8):
            # สร้าง Frame สีขาวขนาด 100x100
            frame = tk.Frame(
                self.frame,
                width=100,
                height=100,
                bg='white',
                relief=tk.RIDGE,
                borderwidth=2
            )
            frame.grid(row=i // 4, column=i % 4, padx=10, pady=10)
            # กำหนดให้คลิกช่องภาพเรียกเมนูนำเข้า
            frame.bind("<Button-1>", lambda event, idx=i: self.show_import_popup(idx))
            
            # สร้างปุ่มลบภาพมุมบนขวา
            btn_delete = tk.Button(
                frame,
                text="X",
                font=("Arial", 8, "bold"),
                fg="white",
                bg="red",
                command=lambda idx=i: self.remove_image(idx)
            )
            btn_delete.place(relx=1, rely=0, anchor='ne')
            
            self.image_boxes.append(frame)
            self.delete_buttons.append(btn_delete)
        
        # ปุ่ม Reset ทุกอย่าง
        self.btn_clear = tk.Button(
            self.root,
            text="Reset",
            bg="#CA5C27",
            fg="white",
            font=("Arial", 12, "bold"),
            relief=tk.RAISED,
            bd=5,
            command=self.clear_all
        )
        self.btn_clear.pack(pady=5)
        
        # ปุ่ม Calculate
        self.btn_calculate = tk.Button(
            self.root,
            text="Calculate",
            bg="green",
            fg="white",
            font=("Arial", 12, "bold"),
            relief=tk.RAISED,
            bd=5,
            command=self.calculate_avg
        )
        self.btn_calculate.pack(pady=5)
        
        # ปุ่ม Save result
        self.btn_save = tk.Button(
            self.root,
            text="Save result",
            bg="lightgray",
            fg="black",
            font=("Arial", 12, "bold"),
            relief=tk.RAISED,
            bd=5,
            command=self.save_results
        )
        self.btn_save.pack(pady=5)
        
        # Label แสดงผลลัพธ์สรุปค่าไขมัน
        self.result_label = tk.Label(
            self.root,
            text="Total AVG fat ____ %",
            font=("Arial", 18, "bold"),
            bg='black',
            fg='white'
        )
        self.result_label.pack(pady=10)
    
    def show_import_popup(self, index):
        # แสดงหน้าต่างย่อยให้เลือกวิธีนำเข้าภาพ
        popup = Toplevel(self.root)
        popup.title("Import Image")
        popup.geometry("200x200")
        
        # ปุ่มอัปโหลดจากไฟล์
        tk.Button(popup, text="Upload Image", bg="yellow", command=lambda: self.add_image(index, popup)).pack(pady=10)
        # ปุ่มถ่ายภาพด้วยกล้อง
        tk.Button(popup, text="Take a Photo", bg="orange", command=lambda: self.capture_image(index, popup)).pack(pady=10)
    
    def add_image(self, index, popup):
        # ฟังก์ชันนำเข้าภาพจากไฟล์
        file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")])
        if file_path:
            image = cv2.imread(file_path)
            self.process_and_display_image(image, index)
        popup.destroy()  # ปิดหน้าต่างย่อย

    def update_datetime(self):
        # อัปเดต Label แสดงวันที่และเวลาเรียลไทม์ทุกวินาที
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.datetime_label.config(text=now)
        if hasattr(self, 'datetime_task'):
            self.root.after_cancel(self.datetime_task)
        self.datetime_task = self.root.after(1000, self.update_datetime)

    def capture_image(self, index, popup):
        # ถ่ายภาพจากกล้องเว็บแคม
        cap = cv2.VideoCapture(0)
        ret, frame = cap.read()
        cap.release()
        if ret:
            self.process_and_display_image(frame, index)
        popup.destroy()
    
    def process_and_display_image(self, image, index):
        # ลบพื้นหลังและคำนวณเปอร์เซ็นต์ไขมัน
        image_processed = image.copy()
        fat_percentage, fat_pixels, total_pixels = self.calculate_fat_percentage(image_processed)

        # แปลงภาพขนาดเล็กสำหรับแสดงใน GUI
        image_display = cv2.resize(image_processed, (100, 100), interpolation=cv2.INTER_AREA)
        image_rgb = cv2.cvtColor(image_display, cv2.COLOR_BGR2RGB)
        image_tk = ImageTk.PhotoImage(Image.fromarray(image_rgb))

        self.clear_existing_image(index)  # ล้างภาพเก่าก่อน

        # วางภาพและเก็บค่าไขมัน
        panel = tk.Label(self.image_boxes[index], image=image_tk)
        panel.image = image_tk
        panel.pack()

        self.image_panels[index] = panel
        self.fat_percentages[index] = fat_percentage
        self.update_fat_label(index, fat_percentage)  # อัปเดต Label แสดงเปอร์เซ็นต์ 
        self.delete_buttons[index].lift()  # ยกปุ่มลบให้อยู่บนสุด

    def calculate_fat_percentage(self, image):
        # สร้าง mask สีขาวและแดงเพื่อแยกไขมันและส่วนเนื้อ
        red_mask = cv2.inRange(image, (0, 0, 60), (150, 150, 255))
        white_mask = cv2.inRange(image, (130, 0, 0), (255, 255, 255))
        fat_pixels = np.count_nonzero(white_mask)
        red_pixels = np.count_nonzero(red_mask)
        white_pixels = np.count_nonzero(white_mask)
        total_pixels = red_pixels + white_pixels
        # คำนวณเปอร์เซ็นต์ไขมัน
        fat_percentage = round((fat_pixels *100) / total_pixels, 2) if total_pixels else 0
        return fat_percentage, fat_pixels, total_pixels

    def update_fat_label(self, index, fat_percentage):
        # ลบ Label เก่าที่มี "fat:"
        self.destroy_widgets_with_text(self.image_boxes[index], "fat:")
        # สร้าง Label ใหม่แสดงเปอร์เซ็นต์ไขมัน
        tk.Label(
            self.image_boxes[index],
            text=f"fat: {fat_percentage}%",
            bg='white',
            font=("Arial", 10, "bold")
        ).pack(side=tk.BOTTOM)

    def remove_image(self, index):
        # ลบภาพและข้อมูลไขมันในช่อง index
        self.clear_existing_image(index)
        self.image_boxes[index].config(bg='white', width=100, height=100)
    
    def clear_existing_image(self, index):
        # ล้าง widget ภาพและ Label "fat:" เดิม
        if self.image_panels[index]:
            self.image_panels[index].destroy()
            self.image_panels[index] = None
        self.fat_percentages[index] = None
        self.destroy_widgets_with_text(self.image_boxes[index], "fat:")
        self.image_boxes[index].config(width=100, height=100, bg='white')

    def clear_all(self):
        # รีเซ็ตทุกช่องภาพและลบผลลัพธ์
        for i in range(8):
            self.remove_image(i)
        self.result_label.config(text="")
    
    def calculate_avg(self):
        # คำนวณค่าเฉลี่ยเปอร์เซ็นต์ไขมันทั้งหมด
        valid_fat_values = [p for p in self.fat_percentages if p is not None]

        if valid_fat_values:
            avg_fat = sum(valid_fat_values) / len(valid_fat_values)

            try:
                target_fat = int(self.target_fat_percentage.get().strip())
                lower_bound = target_fat - 5
                upper_bound = target_fat + 5

                fat_diff = round(avg_fat - target_fat, 2)

                # กำหนดสถานะ Pass/Fail ตามช่วง
                if avg_fat > upper_bound:
                    status = "Fat exceeds the standard"
                elif avg_fat < lower_bound:
                    status = "Fat is below the standard"
                else:
                    status = "Passed"

            except ValueError:
                status = "Invalid target fat input"
                fat_diff = "N/A"

            # แสดงผลลัพธ์ใน GUI
            self.result_label.config(
                text=f"Total AVG fat: {avg_fat:.2f}%\nFat Diff: {fat_diff}%\nStatus: {status}",
                fg="#00FF00" if status == "Passed" else "red",
                font=("Arial", 18, "bold")
            )
            self.calculated_status = True

        else:
            # กรณียังไม่มีข้อมูลภาพ
            self.result_label.config(
                text="No fat data available, please upload images",
                fg="red",
                font=("Arial", 18)
            )
            self.calculated_status = False

    def destroy_widgets_with_text(self, parent, text):
        # ลบ widget Label ที่มีข้อความตรงกับ text
        for widget in parent.winfo_children():
            if isinstance(widget, tk.Label) and text in widget.cget("text"):
                widget.destroy()

    def save_results(self):
        # บันทึกผลลงไฟล์ Excel พร้อมตรวจสอบเงื่อนไขก่อน
        if not self.calculated_status:
            messagebox.showwarning("Warning", "Please calculate before saving.")
            return

        if all(fat is None for fat in self.fat_percentages):
            messagebox.showwarning("Warning", "Images were removed after calculation. Please recalculate before saving.")
            self.result_label.config(text="No fat data available, please upload images", fg="red", font=("Arial", 18))
            return

        if any(image is None for image in self.image_panels):
            messagebox.showwarning("Warning", "Please upload all 8 images before saving.")
            return

        # อ่านค่าจากฟิลด์ ID และ LOT
        id_manufacturer = self.entries["ID_Manufacturer"].get().strip()
        id_qc = self.entries["ID_QC"].get().strip()
        lot_no = self.entries["LOT NO."] .get().strip()

        if not id_manufacturer or not id_qc or not lot_no:
            messagebox.showwarning("Warning", "Please fill in all fields: ID_Manufacturer, ID_QC, and LOT NO.")
            return

        file_path = "results3.xlsx"

        try:
            # เตรียมข้อมูลสำหรับบันทึก
            date_value = self.datetime_label["text"]
            target_fat = int(self.target_fat_percentage.get().strip())

            valid_fats = [p for p in self.fat_percentages if p is not None]
            avg_fat = round(sum(valid_fats) / len(valid_fats), 2)

            lower_bound = target_fat - 5
            upper_bound = target_fat + 5

            if avg_fat > upper_bound:
                status = "Fat exceeds the standard"
            elif avg_fat < lower_bound:
                status = "Fat is below the standard"
            else:
                status = "Passed"

            fat_diff = round(avg_fat - target_fat, 2)
            
        except ValueError:
            messagebox.showerror("Error", "Invalid Target Fat input. Please enter a number between 0–100.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate: {e}")
            return
        
        try:
            # ตรวจสอบและเปิดหรือสร้างไฟล์ Excel
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append([
                    "ID_Manufacturer", "ID_QC", "LOT_NO", "Date_Time",
                    "Target_fat_Percentage", "Avg_fat", "Fat_Diff", "Status",
                    "Img_a1_fat", "Img_a2_fat", "Img_b1_fat", "Img_b2_fat",
                    "Img_c1_fat", "Img_c2_fat", "Img_d1_fat", "Img_d2_fat"
                ])

            # เพิ่มแถวข้อมูลใหม่
            new_data = [
                id_manufacturer, id_qc, lot_no, date_value,
                f"{target_fat}%", f"{avg_fat}%", fat_diff, status
            ]
            ws.append(new_data)
            row_num = ws.max_row

            # ตกแต่งเซลล์ Fat_Diff
            fat_diff_cell = ws["G" + str(row_num)]
            try:
                if abs(fat_diff) > 5:
                    fat_diff_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    fat_diff_cell.font = Font(color="9C0006", bold=True)
                else:
                    fat_diff_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    fat_diff_cell.font = Font(color="006100", bold=True)
            except:
                pass

            # ตกแต่งเซลล์ Status
            status_cell = ws["H" + str(row_num)]
            try:
                if status == "Fat exceeds the standard":
                    status_cell.fill = PatternFill(start_color="FF5353", end_color="FF5353", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                elif status == "Fat is below the standard":
                    status_cell.fill = PatternFill(start_color="FF9B9B", end_color="FF9B9B", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="9AE2A8", end_color="9AE2A8", fill_type="solid")
                    status_cell.font = Font(color="006100", bold=True)
            except: 
                pass

            # แทรกภาพและเปอร์เซ็นต์ไขมันใน Excel
            image_positions = ["I", "J", "K", "L", "M", "N", "O", "P"]

            for i, panel in enumerate(self.image_panels):
                if panel and panel.image:
                    fat_percentage = self.fat_percentages[i] if self.fat_percentages[i] is not None else 0
                    pil_image = ImageTk.getimage(panel.image)
                    img_bytes = BytesIO()
                    pil_image.save(img_bytes, format="PNG")
                    img_bytes.seek(0)

                    excel_img = ExcelImage(img_bytes)
                    image_cell = image_positions[i] + str(row_num)
                    ws.add_image(excel_img, image_cell)

                    # เขียนเปอร์เซ็นต์ไขมันใน cell
                    ws[image_cell] = f"{fat_percentage}%"

            wb.save(file_path)  # บันทึกไฟล์
            messagebox.showinfo("Success", f"Results saved to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    FatDetectionApp(root)
    root.mainloop()


